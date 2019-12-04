from cilantro_audit.completed_audit import CompletedAudit
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


class ExcelFile:
    def __init__(self, title, auditor, dt, dt_obj, **kw):
        super().__init__(**kw)
        self.title = title
        self.auditor = auditor
        self.dt = dt
        self.dt_obj = dt_obj
        self.ca = self.get_completed_audit()

    def get_completed_audit(self):
        ca_list = list(CompletedAudit.objects(datetime=self.dt_obj))
        return ca_list[0]

    def open_file(self, sheet_name, path_to_file):
        wb = Workbook()
        ws = wb.get_active_sheet()
        ws.title = sheet_name
        self.build_sheet(wb, sheet_name)
        return wb

    def build_sheet(self, wb, sheet_name):
        ws = wb.get_sheet_by_name(sheet_name)

        row_counter = 4
        counter = 0

        # Write header to the sheet
        ws.cell(row=1, column=1).value = "Title: " + self.title
        ws.cell(row=1, column=1).font = Font(bold=True)
        ws.cell(row=1, column=2).value = "Auditor: " + self.auditor
        ws.cell(row=1, column=2).font = Font(bold=True)
        ws.cell(row=1, column=3).value = "Date/Time: " + self.dt
        ws.cell(row=1, column=3).font = Font(bold=True)


        # Write row descriptions
        for answer in self.ca.answers:
            ws.cell(row=row_counter, column=1).value = "Question:"
            ws.cell(row=row_counter, column=1).font = Font(bold=True)
            ws.cell(row=row_counter, column=1).alignment = Alignment(vertical='top')
            ws.cell(row=row_counter+1, column=1).value = "Response:"
            ws.cell(row=row_counter+1, column=1).font = Font(bold=True)
            ws.cell(row=row_counter+1, column=1).alignment = Alignment(vertical='top')
            ws.cell(row=row_counter+2, column=1).value = "Comments:"
            ws.cell(row=row_counter+2, column=1).font = Font(bold=True)
            ws.cell(row=row_counter+2, column=1).alignment = Alignment(vertical='top')
            ws.cell(row=row_counter+3, column=1).value = "Severity:"
            ws.cell(row=row_counter+3, column=1).font = Font(bold=True)
            ws.cell(row=row_counter+3, column=1).alignment = Alignment(vertical='top')

            row_counter += 5

        row_counter = 4

        # Write question and answer
        for answer in self.ca.answers:
            ws.cell(row=row_counter, column=2).value = answer.text
            ws.cell(row=row_counter, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row_counter + 1, column=2).value = self.ca.answers[counter].response.response
            ws.cell(row=row_counter + 1, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            if not self.ca.answers[counter].comment:
                ws.cell(row=row_counter + 2, column=2).value = "None"
            else:
                ws.cell(row=row_counter + 2, column=2).value = self.ca.answers[counter].comment
            ws.cell(row=row_counter + 2, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            ws.cell(row=row_counter + 3, column=2).value = self.ca.answers[counter].severity.severity[2:]
            ws.cell(row=row_counter + 3, column=2).alignment = Alignment(wrap_text=True,vertical='top')

            row_counter += 5
            counter += 1

    def print_stuff(self):
        print(self.title, self.auditor, self.dt, self.ca)
